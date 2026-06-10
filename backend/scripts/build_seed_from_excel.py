"""从 Excel 汇总表生成监测源配置（stdout 输出 Python 片段）。"""
from __future__ import annotations

import re
from pathlib import Path
from urllib.parse import urljoin

import openpyxl

EXCEL_PATH = Path(r"d:\下载\全国教育官方网站汇总表.xlsx")

# 已有验证可用的栏目路径（相对域名不变时保留）
COLUMN_OVERRIDES: dict[str, list[dict]] = {
    "北京市教育委员会": [
        {"column_name": "通知公告", "path": "/zwgk/tzgg/", "column_type": "NOTICE", "crawl_interval": 30},
        {"column_name": "政策文件", "path": "/zwgk/zcfg/", "column_type": "POLICY", "crawl_interval": 60},
    ],
    "浙江省教育厅": [
        {"column_name": "通知公告", "path": "/col/col1228996668/index.html", "column_type": "NOTICE", "crawl_interval": 30},
        {"column_name": "政策文件", "path": "/col/col1228996667/index.html", "column_type": "POLICY", "crawl_interval": 60},
        {"column_name": "项目申报", "path": "/col/col1228996670/index.html", "column_type": "PROJECT_APPLY", "crawl_interval": 30},
    ],
    "江苏省教育厅": [
        {"column_name": "通知公告", "path": "/col/col63477/index.html", "column_type": "NOTICE", "crawl_interval": 30},
        {"column_name": "政策解读", "path": "/col/col63478/index.html", "column_type": "POLICY", "crawl_interval": 60},
    ],
    "广东省教育厅": [
        {"column_name": "通知公告", "path": "/zwgk/tzgg/", "column_type": "NOTICE", "crawl_interval": 30},
        {"column_name": "政策法规", "path": "/zwgk/zcfg/", "column_type": "POLICY", "crawl_interval": 60},
    ],
    "山东省教育厅": [
        {"column_name": "通知公告", "path": "/zwgk/tzgg/", "column_type": "NOTICE", "crawl_interval": 30},
        {"column_name": "政策文件", "path": "/zwgk/zcfg/", "column_type": "POLICY", "crawl_interval": 60},
    ],
    "上海市教育委员会": [
        {"column_name": "政务公开", "path": "/cms/web/index/index.jsp", "column_type": "NOTICE", "crawl_interval": 30},
    ],
}

PLAYWRIGHT_SOURCES = {"上海市教育委员会"}

DEFAULT_COLUMNS = [
    {"column_name": "通知公告", "path": "/zwgk/tzgg/", "column_type": "NOTICE", "crawl_interval": 30},
    {"column_name": "政策文件", "path": "/zwgk/zcfg/", "column_type": "POLICY", "crawl_interval": 60},
]

MOE_COLUMNS = [
    {"column_name": "通知公告", "path": "/jyb_xwfb/s271/", "column_type": "NOTICE", "crawl_interval": 30},
    {"column_name": "政策文件", "path": "/jyb_sjzl/s3165/", "column_type": "POLICY", "crawl_interval": 60},
    {"column_name": "项目申报", "path": "/jyb_xwfb/s5984/", "column_type": "PROJECT_APPLY", "crawl_interval": 30},
]


def _load_excel() -> tuple[list, list]:
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    sheets = wb.sheetnames
    national = []
    for r in range(3, wb[sheets[0]].max_row + 1):
        row = [wb[sheets[0]].cell(r, c).value for c in range(1, 5)]
        if row[0]:
            national.append(row)
    provincial = []
    for r in range(3, wb[sheets[1]].max_row + 1):
        row = [wb[sheets[1]].cell(r, c).value for c in range(1, 5)]
        if row[0]:
            provincial.append(row)
    wb.close()
    return national, provincial


def _col_url(base: str, path: str) -> str:
    return urljoin(base.rstrip("/") + "/", path.lstrip("/"))


def _build_source(name: str, url: str, *, moe: bool = False) -> dict:
    base = url.rstrip("/")
    col_defs = MOE_COLUMNS if moe else COLUMN_OVERRIDES.get(name, DEFAULT_COLUMNS)
    columns = []
    for c in col_defs:
        item = dict(c)
        path = item.pop("path")
        item["column_url"] = _col_url(base, path)
        columns.append(item)
    src: dict = {
        "name": name,
        "url": base,
        "type": "SourceType.MOE" if moe else "SourceType.PROVINCIAL",
        "columns": columns,
    }
    if not moe and name in PLAYWRIGHT_SOURCES:
        src["use_playwright"] = True
    return src


def _fmt_columns(columns: list[dict]) -> str:
    lines = []
    for col in columns:
        ctype = f"ColumnType.{col['column_type']}"
        lines.append(
            f'            {{"column_name": "{col["column_name"]}", '
            f'"column_url": "{col["column_url"]}", '
            f'"column_type": {ctype}, '
            f'"crawl_interval": {col["crawl_interval"]}}},'
        )
    return "\n".join(lines)


def _fmt_source(src: dict, *, indent: str = "") -> str:
    extra = ""
    if src.get("use_playwright"):
        extra = f'\n{indent}    "use_playwright": True,'
    cols = _fmt_columns(src["columns"])
    return (
        f'{indent}{{\n'
        f'{indent}    "name": "{src["name"]}",\n'
        f'{indent}    "url": "{src["url"]}",\n'
        f'{indent}    "type": {src["type"]},{extra}\n'
        f'{indent}    "columns": [\n{cols}\n{indent}    ],\n'
        f'{indent}}},'
    )


def main() -> None:
    national, provincial = _load_excel()
    moe_row = national[0]
    moe = _build_source("中华人民共和国教育部", moe_row[2], moe=True)
    prov_sources = [_build_source(row[2], row[3]) for row in provincial]

    out_path = Path(__file__).resolve().parent.parent / "app" / "data" / "monitor_sources_data.py"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    body = (
        '"""监测源预设数据，由 scripts/build_seed_from_excel.py 从 Excel 汇总表生成。"""\n'
        "from app.models.monitor import ColumnType, SourceType\n\n"
        f"MOE_SOURCE = {_fmt_source(moe, indent='').rstrip(',')}\n\n"
        "PROVINCIAL_SOURCES = [\n"
        + "\n".join(_fmt_source(s, indent="    ") for s in prov_sources)
        + "\n]\n"
    )
    out_path.write_text(body, encoding="utf-8")
    print(f"wrote {out_path} ({len(prov_sources)} provincial sources)")


if __name__ == "__main__":
    main()
